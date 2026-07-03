import os

import openpyxl
import pandas as pd
import pytest

import fetch
import process


CR_CSV = """policy_no,company_name,approval_amount,policy_effective_date,policy_renewal_date
POL1,ACME Corp,1000,2026-01-01,2027-01-01
POL2,Beta Co,500,2026-02-01,2027-02-01
POL3,Gamma Co,200,2026-03-01,2027-03-01
"""

BENEFIT_CSV = """policy_no,source_name,approved,nik
POL1,BrokerX,600,0012345678
POL1,BrokerX,400,0098765432
"""

QUERY_CSV = """policy_no,source_name,approved
POL2,BrokerY,500
"""

MASTER = {
    "POL1": {"need_report": "detail"},
    "POL2": {"need_report": "header"},
    "POL3": {"need_report": "detail"},  # no matching Benefit rows -> should be skipped
}


@pytest.fixture
def convert_folder(tmp_path):
    folder = tmp_path / "convert_csv"
    folder.mkdir()
    (folder / fetch.CR_FILENAME).write_text(CR_CSV)
    (folder / fetch.BENEFIT_FILENAME).write_text(BENEFIT_CSV)
    (folder / fetch.QUERY_FILENAME).write_text(QUERY_CSV)
    return str(folder)


def test_process_join_routes_and_skips(tmp_path, convert_folder):
    output_folder = str(tmp_path / "output")

    written = process.process_join(convert_folder, output_folder, master=MASTER, report_period="Juli 2026")

    written_policies = {w["policy_no"] for w in written}
    assert written_policies == {"POL1", "POL2"}  # POL3 skipped: no Benefit rows

    for w in written:
        assert os.path.exists(w["file_path"])

    pol1 = next(w for w in written if w["policy_no"] == "POL1")
    assert pol1["company_name"] == "ACME Corp"
    assert pol1["source_name"] == "BrokerX"
    expected_path = os.path.join(output_folder, "BrokerX", "ACME Corp", "POL1")
    assert os.path.dirname(pol1["file_path"]) == expected_path

    wb = openpyxl.load_workbook(pol1["file_path"])
    assert wb.sheetnames == ["CR", "Benefit"]

    # Leading zeros on nik must survive the round trip (dtype=str on read)
    benefit_ws = wb["Benefit"]
    header = [c.value for c in benefit_ws[1]]
    nik_col = header.index("nik") + 1
    nik_values = [benefit_ws.cell(row=r, column=nik_col).value for r in range(2, benefit_ws.max_row + 1)]
    assert "0012345678" in nik_values

    pol2 = next(w for w in written if w["policy_no"] == "POL2")
    wb2 = openpyxl.load_workbook(pol2["file_path"])
    assert wb2.sheetnames == ["CR", "Query_result"]


def test_process_join_no_cr_file(tmp_path):
    empty_convert = str(tmp_path / "empty_convert")
    os.makedirs(empty_convert)
    output_folder = str(tmp_path / "output")
    written = process.process_join(empty_convert, output_folder, master=MASTER, report_period="Juli 2026")
    assert written == []


def _cr_row(approval_amount):
    return pd.DataFrame([{
        "company_name": "ACME Corp",
        "approval_amount": approval_amount,
        "policy_effective_date": "2026-01-01",
        "policy_renewal_date": "2027-01-01",
    }])


def _data_rows(approved_values, source_name="BrokerX"):
    return pd.DataFrame([
        {"source_name": source_name, "approved": v} for v in approved_values
    ])


def test_write_policy_approval_mismatch_raises(tmp_path):
    df_cr = _cr_row(1000)
    df_data = _data_rows([100, 100])  # sums to 200, not 1000

    with pytest.raises(ValueError, match="approval_amount mismatch"):
        process._write_policy(
            "POL1", df_cr, df_data, str(tmp_path / "output"), "Benefit", "Juli 2026"
        )


def test_write_policy_approval_match_succeeds(tmp_path):
    df_cr = _cr_row(1000)
    df_data = _data_rows([600, 400])  # sums to 1000, matches

    result = process._write_policy(
        "POL1", df_cr, df_data, str(tmp_path / "output"), "Benefit", "Juli 2026"
    )
    assert os.path.exists(result["file_path"])
